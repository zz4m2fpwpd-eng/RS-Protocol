#!/usr/bin/env python3
"""
TITAN Master Excel Builder v1.0
Extracts chart PNG/PDFs into editable Excel sheets per test
Each chart = separate tab with data + metadata/formatting notes
Names master file per test for non-destructive runs
"""

import sys
import json
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os
from PIL import Image
import io


class MasterExcelBuilder:
    """Build editable Excel master with chart data and OCR metadata"""
    
    def __init__(self, test_name, output_dir=None):
        self.test_name = test_name
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = Path(output_dir or "titan_results")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Unique per test - no overwrite
        self.master_xlsx = self.output_dir / f"TITAN_Master_Index_{test_name}.xlsx"
        
        # Setup logging
        self.logger = self._setup_logging()
        self.logger.info(f"Master Excel Builder initialized: {self.master_xlsx}")
        
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Metadata storage
        self.metadata = {
            "test_name": test_name,
            "timestamp": self.timestamp,
            "total_charts": 0,
            "total_sheets": 0,
            "chart_sources": [],  # (chart_file, tab_name)
            "ocr_status": [],
        }
    
    def _setup_logging(self):
        """Setup logging to console + file"""
        logger = logging.getLogger(f"MasterBuilder_{self.test_name}")
        handler = logging.StreamHandler(sys.stdout)
        formatter = logging.Formatter(
            "%(asctime)s | %(name)s | %(levelname)-8s | %(message)s"
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        return logger
    
    def add_index_sheet(self, files_dict):
        """Create INDEX tab: master catalog of all files"""
        self.logger.info("Building INDEX sheet...")
        
        ws = self.wb.create_sheet("INDEX", 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        
        # Headers
        headers = ["Chart/File Name", "Type", "Size (KB)", "Sheet Tab", "Notes"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = cell.font.copy()
            cell.font = cell.font.copy()
            from openpyxl.styles import PatternFill, Font
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row_num = 2
        for filename, file_info in files_dict.items():
            ws[f"A{row_num}"] = filename
            ws[f"B{row_num}"] = file_info.get("type", "unknown")
            size_kb = file_info.get("size_bytes", 0) / 1024
            ws[f"C{row_num}"] = f"{size_kb:.1f}"
            ws[f"D{row_num}"] = file_info.get("sheet_tab", "")
            ws[f"E{row_num}"] = file_info.get("notes", "")
            
            # Link to sheet if available
            if file_info.get("sheet_tab"):
                ws[f"D{row_num}"].hyperlink = f"#{file_info['sheet_tab']}!A1"
                ws[f"D{row_num}"].font = ws[f"D{row_num}"].font.copy()
                from openpyxl.styles import Font
                ws[f"D{row_num}"].font = Font(underline="single", color="0563C1")
            
            row_num += 1
        
        self.metadata["total_sheets"] += 1
        self.logger.info(f"INDEX sheet created with {row_num - 2} entries")
    
    def add_chart_sheet(self, chart_path, sheet_name=None, ocr_data=None):
        """
        Add single chart as a sheet
        - Embed chart image
        - Store extracted OCR data as table
        - Store formatting metadata
        """
        chart_path = Path(chart_path)
        if not chart_path.exists():
            self.logger.warning(f"Chart not found: {chart_path}")
            return False
        
        if sheet_name is None:
            sheet_name = chart_path.stem[:31]  # Max 31 char Excel sheet name
        
        sheet_name = self._sanitize_sheet_name(sheet_name)
        
        try:
            ws = self.wb.create_sheet(sheet_name)
            
            # Row 1: Title + metadata
            ws['A1'] = f"Chart: {chart_path.name}"
            ws['A1'].font = ws['A1'].font.copy()
            from openpyxl.styles import Font
            ws['A1'].font = Font(bold=True, size=12)
            
            ws['A2'] = f"Source File: {chart_path}"
            ws['A3'] = f"Embedded: {datetime.now().isoformat()}"
            ws['A4'] = f"Test: {self.test_name}"
            
            # Embed image starting at row 6
            if chart_path.suffix.lower() in ['.png', '.jpg', '.jpeg']:
                img = XLImage(str(chart_path))
                img.width = 500  # pixels
                img.height = 350
                ws.add_image(img, 'A6')
                current_row = 35  # After image (rough)
            else:
                current_row = 6
            
            # Add OCR/extracted data if available
            if ocr_data:
                ws[f'A{current_row}'] = "Extracted Data"
                ws[f'A{current_row}'].font = Font(bold=True, size=11)
                current_row += 1
                
                # If ocr_data is a DataFrame, write it
                if isinstance(ocr_data, pd.DataFrame):
                    for r_idx, row in enumerate(ocr_data.values, start=current_row):
                        for c_idx, val in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=val)
                
                elif isinstance(ocr_data, dict):
                    # Write as key-value pairs
                    for k, v in ocr_data.items():
                        ws[f'A{current_row}'] = str(k)
                        ws[f'B{current_row}'] = str(v)
                        current_row += 1
            
            # Add formatting notes sheet
            ws[f'A{current_row + 2}'] = "Formatting/Issues"
            ws[f'A{current_row + 2}'].font = Font(bold=True, size=10)
            ws[f'A{current_row + 3}'] = "Chart Type: [Detected type here]"
            ws[f'A{current_row + 4}'] = "Labels Readable: [Yes/No]"
            ws[f'A{current_row + 5}'] = "Data Accuracy: [Check/Edit]"
            ws[f'A{current_row + 6}'] = "Notes: [Add formatting issues]"
            
            self.metadata["total_charts"] += 1
            self.metadata["total_sheets"] += 1
            self.metadata["chart_sources"].append((chart_path.name, sheet_name))
            
            self.logger.info(f"✓ Added chart sheet: {sheet_name}")
            return True
        
        except Exception as e:
            self.logger.error(f"Failed to add chart {chart_path.name}: {e}")
            return False
    
    def add_ocr_results_sheet(self, chart_file_mapping):
        """
        Add OCR_RESULTS tab summarizing all extracted text/data
        chart_file_mapping: {chart_name: extracted_text/data}
        """
        ws = self.wb.create_sheet("OCR_RESULTS")
        
        from openpyxl.styles import PatternFill, Font
        
        headers = ["Chart Name", "OCR Status", "Extracted Text Preview", "Data Points", "Quality"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row_num = 2
        for chart_name, ocr_info in chart_file_mapping.items():
            ws[f"A{row_num}"] = chart_name
            ws[f"B{row_num}"] = ocr_info.get("status", "pending")
            ws[f"C{row_num}"] = str(ocr_info.get("text", ""))[:100]
            ws[f"D{row_num}"] = ocr_info.get("data_points", 0)
            ws[f"E{row_num}"] = ocr_info.get("quality_score", "?")
            row_num += 1
        
        self.metadata["total_sheets"] += 1
        self.logger.info(f"OCR_RESULTS sheet created")
    
    def add_metadata_sheet(self):
        """Add METADATA tab with test info"""
        ws = self.wb.create_sheet("METADATA")
        
        from openpyxl.styles import PatternFill, Font
        
        ws['A1'] = "Test Metadata"
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        for key, value in self.metadata.items():
            ws[f'A{row}'] = str(key)
            ws[f'B{row}'] = str(value)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        
        self.metadata["total_sheets"] += 1
        self.logger.info("METADATA sheet created")
    
    def _sanitize_sheet_name(self, name):
        """Clean sheet name (max 31 chars, no special chars)"""
        # Remove invalid chars
        invalid = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid:
            name = name.replace(char, '_')
        # Truncate to 31
        return name[:31]
    
    def save(self):
        """Save workbook to file"""
        try:
            self.wb.save(str(self.master_xlsx))
            self.logger.info(f"✓ Master Excel saved: {self.master_xlsx}")
            self.logger.info(f"  Sheets: {self.metadata['total_sheets']}")
            self.logger.info(f"  Charts: {self.metadata['total_charts']}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to save: {e}")
            return False


def discover_charts(master_root):
    """Find all chart PNGs/PDFs in MASTER results directory"""
    master_root = Path(master_root)
    
    charts = {}
    for png_file in master_root.glob("**/*.png"):
        charts[png_file.name] = png_file
    for pdf_file in master_root.glob("**/*.pdf"):
        charts[pdf_file.name] = pdf_file
    
    return charts


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="TITAN Master Excel Builder - Convert charts to editable Excel"
    )
    parser.add_argument("--master-root", required=True, help="MASTER results root directory")
    parser.add_argument("--test-name", required=True, help="Test name (for filename)")
    parser.add_argument("--output", default=".", help="Output directory for Excel file")
    
    args = parser.parse_args()
    
    builder = MasterExcelBuilder(args.test_name, args.output)
    
    # Find all charts
    charts = discover_charts(args.master_root)
    print(f"Found {len(charts)} charts to process")
    
    # Add INDEX sheet first
    files_dict = {}
    for chart_name, chart_path in charts.items():
        size_bytes = chart_path.stat().st_size
        files_dict[chart_name] = {
            "type": chart_path.suffix.upper(),
            "size_bytes": size_bytes,
            "sheet_tab": chart_path.stem[:31],
            "notes": "See chart sheet for embedded image + data"
        }
    
    builder.add_index_sheet(files_dict)
    
    # Add chart sheets
    for chart_name, chart_path in sorted(charts.items()):
        builder.add_chart_sheet(chart_path, sheet_name=chart_path.stem)
    
    # Add OCR results summary
    ocr_mapping = {
        chart_name: {
            "status": "embedded",
            "text": "",
            "data_points": 0,
            "quality_score": "TBD"
        }
        for chart_name in charts.keys()
    }
    builder.add_ocr_results_sheet(ocr_mapping)
    
    # Add metadata
    builder.add_metadata_sheet()
    
    # Save
    builder.save()
    print(f"\n✓ Master Excel created: {builder.master_xlsx}")
