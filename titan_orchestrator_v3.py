#!/usr/bin/env python3
"""
TITAN RS UNIVERSAL ORCHESTRATOR v3.0 - WITH MASTER EXCEL BUILDER
Parallel execution + automatic Master Excel generation per test
Each test gets unique master file (no overwrite)
Developed by Robin Sandhu
"""

import sys
import json
import logging
import argparse
import subprocess
import multiprocessing
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from concurrent.futures import ProcessPoolExecutor, as_completed
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font
import os


@dataclass
class Config:
    """Central configuration for TITAN orchestrator"""
    PROJECT_NAME = "TITAN-RS"
    VERSION = "3.0"
    TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    ENGINES = {
        "omni": "TITAN_Omni_Protocol.py",
        "research": "TITAN_Research_Mode.py",
        "results": "TITAN_Results_Engine.py",
        "fork": "TITAN_RS_Fork.py",
        "evidence": "TITAN_Evidence_Pro_Max.py",
        "rstitan": "RSTITAN.py",
    }
    
    RESULT_SUBDIRS = ["charts", "reports", "data", "logs", "metadata", "xlsxoutput"]


def setup_logging(logdir):
    """Setup console + file logging"""
    logdir = Path(logdir)
    logdir.mkdir(parents=True, exist_ok=True)
    logfile = logdir / f"titan_run_{Config.TIMESTAMP}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
        handlers=[
            logging.FileHandler(logfile),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)


class InputDiscovery:
    """Discover and validate input files/folders"""
    ALLOWED_FORMATS = [".csv", ".xlsx", ".xls", ".data", ".json", ".parquet", ".pkl"]
    
    @staticmethod
    def discover_files(path_or_pattern, recursive=False):
        """Find all data files in given path/pattern"""
        path = Path(path_or_pattern).expanduser()
        
        if not path.exists():
            raise FileNotFoundError(f"Path not found: {path}")
        
        files_found = []
        
        if path.is_file():
            if path.suffix in InputDiscovery.ALLOWED_FORMATS:
                files_found.append(path)
            else:
                raise ValueError(f"Unsupported file format: {path.suffix}")
        
        elif path.is_dir():
            glob_pattern = "**/*" if recursive else "*"
            for f in path.glob(glob_pattern):
                if f.is_file() and f.suffix in InputDiscovery.ALLOWED_FORMATS:
                    files_found.append(f)
        
        return sorted(files_found)
    
    @staticmethod
    def validate_files(files, logger):
        """Validate each file exists and is readable"""
        valid = []
        for f in files:
            try:
                size_mb = f.stat().st_size / 1024 / 1024
                logger.info(f"✓ Found: {f.name} ({size_mb:.2f} MB)")
                valid.append(f)
            except Exception as e:
                logger.warning(f"Skipped {f.name} - {str(e)}")
        return valid


class ResultOrganizer:
    """Organize and manage output files"""
    
    def __init__(self, base_output_dir, test_name):
        """Initialize result directory structure"""
        self.base = Path(base_output_dir).expanduser()
        self.testname = test_name
        self.run_timestamp = Config.TIMESTAMP
        
        self.root = self.base / f"{test_name}_{self.run_timestamp}"
        self.root.mkdir(parents=True, exist_ok=True)
        
        # Create subdirectories
        self.subdirs = {}
        for subdir in Config.RESULT_SUBDIRS:
            subpath = self.root / subdir
            subpath.mkdir(exist_ok=True)
            self.subdirs[subdir] = subpath
        
        # Metadata tracking
        self.metadata = {
            "test_name": test_name,
            "timestamp": self.run_timestamp,
            "engines_used": [],
            "input_files": [],
            "status": "initialized",
            "errors": [],
            "charts_generated": 0,
            "data_files_generated": 0,
            "xlsx_files_generated": 0,
        }
    
    def register_input(self, filepath):
        """Register an input file"""
        self.metadata["input_files"].append(str(filepath))
    
    def move_charts(self, source_dir, logger):
        """Find and move all chart files"""
        chart_exts = [".png", ".jpg", ".jpeg", ".pdf", ".svg"]
        count = 0
        source = Path(source_dir)
        
        if not source.exists():
            logger.warning(f"Chart source dir not found: {source}")
            return count
        
        for chart_file in source.glob("*"):
            if chart_file.suffix.lower() in chart_exts:
                try:
                    dest = self.subdirs["charts"] / chart_file.name
                    chart_file.rename(dest)
                    count += 1
                except Exception as e:
                    logger.warning(f"Could not move chart {chart_file.name}: {e}")
        
        self.metadata["charts_generated"] += count
        return count
    
    def move_xlsx(self, source_dir, logger):
        """Find and move all Excel files"""
        xlsx_exts = [".xlsx", ".xls"]
        count = 0
        source = Path(source_dir)
        
        if not source.exists():
            logger.warning(f"XLSX source dir not found: {source}")
            return count
        
        for xlsx_file in source.glob("*"):
            if xlsx_file.suffix.lower() in xlsx_exts:
                try:
                    dest = self.subdirs["xlsxoutput"] / xlsx_file.name
                    xlsx_file.rename(dest)
                    count += 1
                except Exception as e:
                    logger.warning(f"Could not move Excel {xlsx_file.name}: {e}")
        
        self.metadata["xlsx_files_generated"] += count
        return count
    
    def move_data_files(self, source_dir, logger):
        """Find and move all result data files"""
        data_exts = [".csv", ".json", ".parquet"]
        count = 0
        source = Path(source_dir)
        
        if not source.exists():
            logger.warning(f"Data source dir not found: {source}")
            return count
        
        for data_file in source.glob("*"):
            if data_file.suffix.lower() in data_exts:
                try:
                    dest = self.subdirs["data"] / data_file.name
                    data_file.rename(dest)
                    count += 1
                except Exception as e:
                    logger.warning(f"Could not move data {data_file.name}: {e}")
        
        self.metadata["data_files_generated"] += count
        return count
    
    def save_metadata(self):
        """Save run metadata as JSON"""
        metafile = self.root / "metadata.json"
        with open(metafile, "w") as f:
            json.dump(self.metadata, f, indent=2)
    
    def save_manifest(self):
        """Generate file manifest"""
        manifest_file = self.root / "MANIFEST.txt"
        with open(manifest_file, "w") as f:
            f.write("=" * 80 + "\n")
            f.write("TITAN RS Test Results Manifest\n")
            f.write(f"Test Name: {self.metadata['test_name']}\n")
            f.write(f"Timestamp: {self.metadata['timestamp']}\n")
            f.write(f"Status: {self.metadata['status']}\n")
            f.write("=" * 80 + "\n")
            
            f.write("\nINPUT FILES:\n")
            for inp in self.metadata["input_files"]:
                f.write(f"  - {inp}\n")
            f.write(f"  ({len(self.metadata['input_files'])} inputs)\n")
            
            f.write("\nOUTPUT STRUCTURE:\n")
            for name, subdir in self.subdirs.items():
                files = list(subdir.glob("*"))
                f.write(f"  {name}: {len(files)} files\n")
                for fpath in sorted(files)[:10]:
                    f.write(f"    - {fpath.name}\n")
                if len(files) > 10:
                    f.write(f"    ... and {len(files) - 10} more\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("SUMMARY\n")
            f.write(f"  Charts: {self.metadata['charts_generated']}\n")
            f.write(f"  XLSX files: {self.metadata['xlsx_files_generated']}\n")
            f.write(f"  Data files: {self.metadata['data_files_generated']}\n")
            f.write("=" * 80 + "\n")


class MasterExcelBuilder:
    """Build editable Excel master with charts as embedded tabs"""
    
    def __init__(self, test_name, output_dir, logger):
        self.test_name = test_name
        self.output_dir = Path(output_dir)
        self.logger = logger
        
        # Unique per test name - no overwrite
        self.master_xlsx = self.output_dir / f"TITAN_Master_Index_{test_name}.xlsx"
        
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        self.metadata = {
            "test_name": test_name,
            "timestamp": datetime.now().isoformat(),
            "total_charts": 0,
            "total_sheets": 0,
            "chart_sources": [],
        }
    
    def add_index_sheet(self, charts_dict):
        """Create INDEX tab"""
        self.logger.info("Building INDEX sheet...")
        
        ws = self.wb.create_sheet("INDEX", 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        
        headers = ["Chart Name", "Type", "Size (KB)", "Sheet Tab", "Notes"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row_num = 2
        for chart_name, chart_path in sorted(charts_dict.items()):
            size_kb = chart_path.stat().st_size / 1024
            sheet_tab = chart_path.stem[:31]
            
            ws[f"A{row_num}"] = chart_name
            ws[f"B{row_num}"] = chart_path.suffix.upper()
            ws[f"C{row_num}"] = f"{size_kb:.1f}"
            ws[f"D{row_num}"] = sheet_tab
            ws[f"E{row_num}"] = "Click to jump to chart"
            
            # Hyperlink to sheet
            ws[f"D{row_num}"].hyperlink = f"#{sheet_tab}!A1"
            ws[f"D{row_num}"].font = Font(underline="single", color="0563C1")
            
            row_num += 1
        
        self.metadata["total_sheets"] += 1
    
    def add_chart_sheet(self, chart_path, sheet_name=None):
        """Add chart as sheet with embedded image"""
        chart_path = Path(chart_path)
        
        if sheet_name is None:
            sheet_name = chart_path.stem[:31]
        
        sheet_name = self._sanitize_sheet_name(sheet_name)
        
        try:
            ws = self.wb.create_sheet(sheet_name)
            
            # Metadata header
            ws['A1'] = f"Chart: {chart_path.name}"
            ws['A1'].font = Font(bold=True, size=12)
            ws['A2'] = f"File: {chart_path}"
            ws['A3'] = f"Embedded: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"Test: {self.test_name}"
            
            # Embed image at row 6
            if chart_path.suffix.lower() in ['.png', '.jpg', '.jpeg']:
                try:
                    img = XLImage(str(chart_path))
                    img.width = 550
                    img.height = 380
                    ws.add_image(img, 'A6')
                    self.logger.info(f"  Embedded image: {chart_path.name}")
                except Exception as e:
                    ws['A35'] = f"Could not embed image: {e}"
                    self.logger.warning(f"  Image embed failed: {e}")
            else:
                ws['A6'] = f"PDF/Vector file not embeddable in Excel: {chart_path.name}"
            
            # Add data extraction area (for future OCR)
            ws['A38'] = "Extracted Data / OCR Results"
            ws['A38'].font = Font(bold=True, size=11)
            ws['A39'] = "Chart Type:"
            ws['B39'] = "[Manually detect or OCR]"
            ws['A40'] = "Labels Readable:"
            ws['B40'] = "[Yes/No - Check embedding]"
            ws['A41'] = "Data Points:"
            ws['B41'] = "[Count from chart]"
            ws['A42'] = "Formatting Notes:"
            ws['B42'] = "[Edit spacing, colors, text issues]"
            
            # Add reference link back to INDEX
            ws['A50'] = "← Back to INDEX"
            ws['A50'].hyperlink = "#INDEX!A1"
            ws['A50'].font = Font(underline="single", color="0563C1")
            
            self.metadata["total_charts"] += 1
            self.metadata["total_sheets"] += 1
            self.metadata["chart_sources"].append((chart_path.name, sheet_name))
            
            return True
        
        except Exception as e:
            self.logger.error(f"Failed to add chart sheet {chart_path.name}: {e}")
            return False
    
    def add_metadata_sheet(self):
        """Add METADATA sheet"""
        ws = self.wb.create_sheet("METADATA")
        
        ws['A1'] = "Test Metadata"
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        for key, value in self.metadata.items():
            ws[f'A{row}'] = str(key)
            ws[f'B{row}'] = str(value)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        
        self.metadata["total_sheets"] += 1
    
    def _sanitize_sheet_name(self, name):
        """Clean sheet name (max 31 chars, no special chars)"""
        invalid = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid:
            name = name.replace(char, '_')
        return name[:31]
    
    def save(self):
        """Save workbook"""
        try:
            self.wb.save(str(self.master_xlsx))
            self.logger.info(f"✓ Master Excel saved: {self.master_xlsx}")
            self.logger.info(f"  Total sheets: {self.metadata['total_sheets']}")
            self.logger.info(f"  Total charts embedded: {self.metadata['total_charts']}")
            return self.master_xlsx
        except Exception as e:
            self.logger.error(f"Failed to save master Excel: {e}")
            return None


class EngineRunner:
    """Execute chosen TITAN engine with error handling"""
    
    def __init__(self, engine_name, logger):
        self.engine_name = engine_name.lower()
        self.logger = logger
        self.engine_file = Config.ENGINES.get(self.engine_name)
        
        if not self.engine_file:
            raise ValueError(
                f"Unknown engine: {self.engine_name}. "
                f"Available: {list(Config.ENGINES.keys())}"
            )
    
    def validate_engine(self):
        """Check if engine file exists"""
        engine_path = Path(self.engine_file)
        if not engine_path.exists():
            self.logger.error(f"Engine file not found: {self.engine_file}")
            return False
        self.logger.info(f"✓ Engine found: {self.engine_name} ({self.engine_file})")
        return True
    
    def run(self, input_file, output_dir):
        """Execute engine on input file"""
        self.logger.info(f"Running engine: {self.engine_name}")
        self.logger.info(f"Input: {input_file.name}")
        self.logger.info(f"Output dir: {output_dir}")
        
        try:
            cmd = [
                sys.executable,
                self.engine_file,
                "--input", str(input_file),
                "--output", str(output_dir),
            ]
            self.logger.info(f"Command: {' '.join(cmd)}")
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=3600,
                stdin=subprocess.DEVNULL,
            )
            
            if result.stdout.strip():
                self.logger.info(f"[STDOUT] {result.stdout[:2000]}")
            
            if result.stderr.strip():
                self.logger.error(f"[STDERR] {result.stderr[:2000]}")
            
            output_log = Path(output_dir) / f"{self.engine_name}_engine_output.log"
            with open(output_log, "w") as f:
                f.write(f"Command: {' '.join(cmd)}\n")
                f.write(f"Return Code: {result.returncode}\n\n")
                f.write(f"STDOUT:\n{result.stdout}\n\n")
                f.write(f"STDERR:\n{result.stderr}\n")
            
            if result.returncode == 0:
                self.logger.info("✓ Completed successfully")
                return True
            else:
                self.logger.error(f"Engine failed with code {result.returncode}")
                return False
        
        except subprocess.TimeoutExpired:
            self.logger.error("Engine execution timed out (1 hour)")
            return False
        except Exception as e:
            self.logger.error(f"Execution error: {str(e)}\n{traceback.format_exc()}")
            return False


def run_engine_on_file(engine_name, input_file, output_dir, test_name, logger):
    """Worker function: run single engine on single file"""
    try:
        runner = EngineRunner(engine_name, logger)
        if not runner.validate_engine():
            return False
        return runner.run(input_file, output_dir)
    except Exception as e:
        logger.error(f"Worker error ({engine_name}): {str(e)}")
        return False


class TitanOrchestratorMulticore:
    """Main orchestrator with Master Excel builder"""
    
    def __init__(self):
        self.logger = None
        self.organizer = None
        self.files = []
    
    def run(self, args):
        """Main execution flow"""
        try:
            print("=" * 80)
            print("TITAN RS UNIVERSAL ORCHESTRATOR v3.0 - MULTICORE + MASTER EXCEL")
            print("=" * 80)
            
            self.organizer = ResultOrganizer(args.output_dir, args.test_name)
            self.logger = setup_logging(self.organizer.subdirs["logs"])
            
            self.logger.info(f"Initialized test: {args.test_name}")
            self.logger.info(f"Result root: {self.organizer.root}")
            
            # --- STEP 1: INPUT DISCOVERY ---
            self.logger.info("\n--- STEP 1: INPUT DISCOVERY ---")
            self.files = InputDiscovery.discover_files(args.input, recursive=args.recursive)
            self.logger.info(f"Found {len(self.files)} files")
            
            if not self.files:
                raise ValueError("No input files found!")
            
            valid_files = InputDiscovery.validate_files(self.files, self.logger)
            self.logger.info(f"Validated {len(valid_files)} files")
            
            # --- STEP 2: ENGINE SELECTION ---
            self.logger.info("\n--- STEP 2: ENGINE SELECTION ---")
            
            if args.engine.lower() == "auto":
                print("\nAvailable TITAN RS Engines:")
                engines = list(Config.ENGINES.keys())
                for i, eng in enumerate(engines, 1):
                    print(f"  {i}. {eng}")
                while True:
                    try:
                        choice = input(f"Select engine(s) (1-{len(engines)}, comma-separated, or 'all'): ").strip()
                        if choice.lower() == "all":
                            engines_to_run = engines
                            break
                        choices = [int(c.strip()) - 1 for c in choice.split(",")]
                        if all(0 <= idx < len(engines) for idx in choices):
                            engines_to_run = [engines[idx] for idx in choices]
                            break
                    except Exception:
                        pass
                    print(f"Invalid choice, try again.")
            else:
                engines_to_run = args.engine.lower().split(",")
            
            self.logger.info(f"Will run {len(engines_to_run)} engine(s): {', '.join(engines_to_run)}")
            
            # --- STEP 3: PARALLEL PROCESSING ---
            self.logger.info(f"\n--- STEP 3: PARALLEL PROCESSING ---")
            self.logger.info(f"Running {len(engines_to_run)} engines × {len(valid_files)} files in parallel")
            
            max_workers = max(1, multiprocessing.cpu_count() - 1)
            self.logger.info(f"Using {max_workers} CPU cores")
            
            tasks = []
            for engine_name in engines_to_run:
                for input_file in valid_files:
                    self.organizer.register_input(input_file)
                    tasks.append((engine_name, input_file))
            
            success_count = 0
            error_count = 0
            
            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                future_to_task = {
                    executor.submit(run_engine_on_file, engine, f, self.organizer.root, args.test_name, self.logger): (engine, f)
                    for engine, f in tasks
                }
                
                for future in as_completed(future_to_task):
                    engine, input_file = future_to_task[future]
                    try:
                        result = future.result()
                        if result:
                            success_count += 1
                            self.logger.info(f"✓ {engine} on {input_file.name} completed")
                        else:
                            error_count += 1
                            self.organizer.metadata["errors"].append(
                                f"{engine} on {input_file.name} failed"
                            )
                    except Exception as e:
                        error_count += 1
                        self.organizer.metadata["errors"].append(
                            f"{engine} on {input_file.name}: {str(e)}"
                        )
                        self.logger.error(f"✗ {engine} on {input_file.name}: {str(e)}")
            
            # --- STEP 4: RESULT ORGANIZATION ---
            self.logger.info(f"\n--- STEP 4: RESULT ORGANIZATION ---")
            charts_moved = self.organizer.move_charts(self.organizer.root, self.logger)
            xlsx_moved = self.organizer.move_xlsx(self.organizer.root, self.logger)
            data_moved = self.organizer.move_data_files(self.organizer.root, self.logger)
            self.logger.info(
                f"Organized {charts_moved} charts, {xlsx_moved} xlsx, {data_moved} data files"
            )
            
            # --- STEP 5: MASTER EXCEL GENERATION ---
            self.logger.info(f"\n--- STEP 5: MASTER EXCEL GENERATION ---")
            master_builder = MasterExcelBuilder(args.test_name, self.organizer.base, self.logger)
            
            # Discover all charts in charts subdir
            charts_dir = self.organizer.subdirs["charts"]
            charts_dict = {}
            for chart_file in sorted(charts_dir.glob("*")):
                if chart_file.suffix.lower() in ['.png', '.jpg', '.jpeg', '.pdf']:
                    charts_dict[chart_file.name] = chart_file
            
            if charts_dict:
                self.logger.info(f"Found {len(charts_dict)} charts to embed")
                master_builder.add_index_sheet(charts_dict)
                
                for chart_name, chart_path in sorted(charts_dict.items()):
                    master_builder.add_chart_sheet(chart_path)
                
                master_builder.add_metadata_sheet()
                excel_file = master_builder.save()
                
                if excel_file:
                    self.logger.info(f"Master Excel: {excel_file}")
            else:
                self.logger.info("No charts found to embed in master Excel")
            
            # --- STEP 6: FINALIZATION ---
            self.logger.info(f"\n--- STEP 6: FINALIZATION ---")
            self.organizer.metadata["status"] = "completed" if error_count == 0 else "completed_with_errors"
            self.organizer.save_metadata()
            self.organizer.save_manifest()
            
            self.logger.info("=" * 80)
            self.logger.info("SUMMARY")
            self.logger.info(f"Tasks executed: {success_count + error_count} ({success_count} success, {error_count} errors)")
            self.logger.info(f"Results saved to: {self.organizer.root}")
            self.logger.info(f"Master Excel: {self.organizer.base / f'TITAN_Master_Index_{args.test_name}.xlsx'}")
            self.logger.info("=" * 80)
            
            print(f"\n✓ All complete! Results in {self.organizer.root}")
            return 0 if error_count == 0 else 1
        
        except Exception as e:
            if self.logger:
                self.logger.error(f"FATAL: {str(e)}\n{traceback.format_exc()}")
            else:
                print(f"FATAL ERROR: {str(e)}")
                traceback.print_exc()
            return 2


def main():
    parser = argparse.ArgumentParser(
        description="TITAN RS Universal Orchestrator v3.0 - Multicore + Master Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 titan_orchestrator_multicore_v3.py --input data.csv --engine omni --test-name heart_2022_no_nans
  python3 titan_orchestrator_multicore_v3.py --input data/ --engine all --test-name full_suite
  python3 titan_orchestrator_multicore_v3.py --input data/ --engine "omni,rstitan" --recursive
        """,
    )
    
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Input file or folder path",
    )
    parser.add_argument(
        "--engine", "-e",
        default="auto",
        help="TITAN engine(s) to use (omni, rstitan, research, results, fork, evidence). Default: auto",
    )
    parser.add_argument(
        "--output-dir", "-o",
        default="titan_results",
        help="Base output directory (default: titan_results)",
    )
    parser.add_argument(
        "--test-name", "-t",
        default=f"titan_test_{datetime.now().strftime('%Y%m%d_%H%M')}",
        help="Test name for result folder + master Excel naming",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Recursively search input directory",
    )
    
    args = parser.parse_args()
    
    orch = TitanOrchestratorMulticore()
    return orch.run(args)


if __name__ == "__main__":
    if sys.platform == "darwin":
        multiprocessing.set_start_method("spawn", force=True)
    multiprocessing.freeze_support()
    
    sys.exit(main())
